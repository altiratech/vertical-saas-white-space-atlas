from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import math
import re
import urllib.request
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
CLEAN_DIR = ROOT / "clean"
DATA_DIR = ROOT / "data"
SITE_DIR = ROOT / "site"

USER_AGENT = (
    "Mozilla/5.0 (compatible; VerticalSaaSWhiteSpaceAtlas/0.1; "
    "+https://local.vertical-saas-white-space-atlas)"
)

CBP_URL = (
    "https://api.census.gov/data/2022/cbp"
    "?get=NAICS2017,NAICS2017_LABEL,ESTAB,EMP,PAYANN&for=us:1&NAICS2017=*"
)
SBA_URL = (
    "https://www.sba.gov/sites/default/files/2023-06/"
    "Table%20of%20Size%20Standards_Effective%20March%2017%2C%202023_.xlsx"
)
BLS_URL = "https://data.bls.gov/cew/data/api/2024/a/area/US000.csv"
NAICS_CROSSWALK_URL = (
    "https://www.census.gov/naics/concordances/2022_to_2017_NAICS_Changes_Only.xlsx"
)
BLS_MATRIX_HOME_URL = "https://data.bls.gov/projections/nationalMatrixHome?ioType=i"
BLS_MATRIX_QUERY_URL_TEMPLATE = "https://data.bls.gov/projections/nationalMatrix?queryParams={code}&ioType=i"
ONET_OCCUPATION_DATA_URL = (
    "https://www.onetcenter.org/dl_files/database/db_30_2_text/Occupation%20Data.txt"
)
ONET_WORK_ACTIVITIES_URL = (
    "https://www.onetcenter.org/dl_files/database/db_30_2_text/Work%20Activities.txt"
)

CBP_RAW_PATH = RAW_DIR / "cbp_2022_us_naics2017.json"
SBA_RAW_PATH = RAW_DIR / "sba_size_standards_2023.xlsx"
BLS_RAW_PATH = RAW_DIR / "bls_qcew_2024_us000_annual_area.csv"
NAICS_CROSSWALK_RAW_PATH = RAW_DIR / "naics_2022_to_2017_changes_only.xlsx"
BLS_MATRIX_HOME_RAW_PATH = RAW_DIR / "bls_nem_2024_2034_industry_home.html"
BLS_MATRIX_PROFILES_RAW_PATH = RAW_DIR / "bls_nem_2024_2034_industry_profiles.json"
ONET_OCCUPATION_DATA_RAW_PATH = RAW_DIR / "onet_30_2_occupation_data.txt"
ONET_WORK_ACTIVITIES_RAW_PATH = RAW_DIR / "onet_30_2_work_activities.txt"

CLEAN_CBP_2017_PATH = CLEAN_DIR / "cbp_2022_national_naics2017.csv"
CLEAN_CBP_2022_PATH = CLEAN_DIR / "cbp_2022_national_naics2022.csv"
CLEAN_BLS_PATH = CLEAN_DIR / "bls_qcew_2024_national_private_naics2022.csv"
CLEAN_SBA_PATH = CLEAN_DIR / "sba_size_standards_2023_naics2022.csv"
CLEAN_CROSSWALK_PATH = CLEAN_DIR / "naics_2022_to_2017_crosswalk.csv"
CLEAN_INDUSTRY_TABLE_PATH = CLEAN_DIR / "industry_table_national_naics2022.csv"
CLEAN_BLS_MATRIX_DIRECTORY_PATH = CLEAN_DIR / "bls_nem_2024_2034_industry_directory.csv"
CLEAN_WORKFLOW_PROFILE_PATH = CLEAN_DIR / "industry_workflow_profiles_national_naics2022.csv"

DATA_CELLS_JSON_PATH = DATA_DIR / "industry_cells.json"
DATA_CELLS_CSV_PATH = DATA_DIR / "industry_cells.csv"
DATA_COVERAGE_GAPS_PATH = DATA_DIR / "coverage_gaps.csv"
SITE_DATA_PATH = SITE_DIR / "data.json"

SOURCE_ARTIFACT_SCHEMA_PATH = ROOT / "schemas" / "source_artifact.schema.json"
SITE_PAYLOAD_SCHEMA_PATH = ROOT / "schemas" / "site_payload.schema.json"

NUMERIC_CODE_RE = re.compile(r"\d{6}")
SOC_CODE_RE = re.compile(r"\d{2}-\d{4}")

WORKFLOW_COMPONENT_WEIGHTS: dict[str, list[tuple[str, float]]] = {
    "documentation": [
        ("Documenting/Recording Information", 0.45),
        ("Processing Information", 0.3),
        ("Performing Administrative Activities", 0.25),
    ],
    "coordination": [
        ("Scheduling Work and Activities", 0.35),
        ("Communicating with Supervisors, Peers, or Subordinates", 0.35),
        ("Coordinating the Work and Activities of Others", 0.3),
    ],
    "compliance": [
        ("Evaluating Information to Determine Compliance with Standards", 0.35),
        ("Monitoring Processes, Materials, or Surroundings", 0.25),
        ("Updating and Using Relevant Knowledge", 0.2),
        ("Inspecting Equipment, Structures, or Materials", 0.2),
    ],
    "care_service": [
        ("Assisting and Caring for Others", 1.0),
    ],
}
WORKFLOW_COMPONENT_LABELS = {
    "documentation": "documentation",
    "coordination": "coordination",
    "compliance": "compliance",
    "care_service": "care and service",
}
WORKFLOW_SCORE_WEIGHTS = {
    "documentation": 0.3,
    "coordination": 0.3,
    "compliance": 0.25,
    "care_service": 0.15,
}
FRONTLINE_OPERATOR_MAJOR_GROUPS = {
    "29",
    "31",
    "33",
    "35",
    "37",
    "39",
    "41",
    "43",
    "47",
    "49",
    "51",
    "53",
}
KNOWLEDGE_WORK_MAJOR_GROUPS = {"13", "15", "17", "19", "21", "23", "25", "27"}

SECTOR_FIT_ADJUSTMENTS: dict[str, tuple[float, str]] = {
    "23": (10.0, "Construction and trade workflows tend to create recurring operator software needs."),
    "31": (3.0, "Physical operations can support repeatable workflow software, but are not the clearest first wedge."),
    "32": (3.0, "Physical operations can support repeatable workflow software, but are not the clearest first wedge."),
    "33": (3.0, "Physical operations can support repeatable workflow software, but are not the clearest first wedge."),
    "42": (5.0, "Merchant distribution workflows can create real coordination pain."),
    "44": (4.0, "Multi-site retail operations can create recurring coordination and labor workflows."),
    "45": (4.0, "Multi-site retail operations can create recurring coordination and labor workflows."),
    "48": (10.0, "Logistics and routing workflows fit the operator-heavy wedge thesis."),
    "49": (10.0, "Storage and delivery workflows fit the operator-heavy wedge thesis."),
    "53": (3.0, "Property operations can be viable, but buyer quality varies widely."),
    "56": (12.0, "Outsourced operator services often have repeated labor and compliance workflows."),
    "62": (10.0, "Care delivery workflows are operations-heavy and often documentation-heavy."),
    "72": (12.0, "Hospitality and food-service markets are location-heavy, labor-heavy, and workflow-driven."),
    "81": (10.0, "Local recurring service operations are often good vertical-software terrain."),
    "22": (-16.0, "Utility markets are capital-intensive and often less practical as a first software wedge."),
    "51": (-22.0, "Information and software-native markets are less aligned with the first-wedge thesis."),
    "52": (-22.0, "Finance and insurance markets are less aligned with the first-wedge thesis."),
    "54": (-6.0, "Specialist professional services can skew toward knowledge-work rather than repeatable operator workflows."),
    "55": (-12.0, "Holding-company structures are not useful first-wedge markets."),
    "61": (-10.0, "Education buyers are often institution-heavy and slower to wedge."),
    "71": (-14.0, "Arts, entertainment, and recreation are less consistent with the intended first-wedge profile."),
}

POSITIVE_TITLE_SIGNALS: list[tuple[str, float, str]] = [
    ("food service", 16.0, "Food-service operations create recurring staffing, scheduling, and unit-level execution work."),
    ("hotel", 14.0, "Hotel operations create recurring property, staffing, and guest-service workflows."),
    ("motel", 14.0, "Lodging operations create recurring property, staffing, and guest-service workflows."),
    ("home health", 16.0, "Home-health operations are distributed, labor-heavy, and documentation-heavy."),
    ("assisted living", 14.0, "Assisted-living workflows combine care delivery, staffing, and compliance."),
    ("nursing care", 14.0, "Nursing-care workflows combine care delivery, staffing, and compliance."),
    ("outpatient", 10.0, "Outpatient operations often depend on repeatable scheduling and documentation workflows."),
    ("dialysis", 12.0, "Dialysis operations are repeated, regulated, and workflow-heavy."),
    ("laborator", 10.0, "Laboratory workflows are operationally repeatable and process-heavy."),
    ("warehouse", 14.0, "Warehousing operations create repeated labor, routing, and throughput workflows."),
    ("warehousing", 14.0, "Warehousing operations create repeated labor, routing, and throughput workflows."),
    ("courier", 14.0, "Courier workflows are routing-heavy and operationally repetitive."),
    ("freight", 12.0, "Freight workflows are routing-heavy and operationally repetitive."),
    ("parking", 14.0, "Parking operations are recurring, local, and workflow-driven."),
    ("solid waste", 14.0, "Waste operations are route-based, labor-heavy, and recurring."),
    ("waste collection", 14.0, "Waste operations are route-based, labor-heavy, and recurring."),
    ("security guards", 14.0, "Security operations create recurring staffing, routing, and compliance workflows."),
    ("patrol", 12.0, "Patrol operations create recurring staffing, routing, and compliance workflows."),
    ("janitorial", 14.0, "Janitorial operations create recurring staffing and multi-site service workflows."),
    ("facilities support", 14.0, "Facilities support contracts create recurring service workflows."),
    ("temporary help", 12.0, "Staffing markets create recurring assignment, scheduling, and communication loops."),
    ("grocery", 10.0, "Grocery operations combine multi-site labor, inventory, and local execution workflows."),
    ("supermarket", 10.0, "Supermarket operations combine multi-site labor, inventory, and local execution workflows."),
    ("gasoline stations", 8.0, "Convenience-retail operations create repeated local labor and site workflows."),
    ("contractor", 14.0, "Contractor markets create repeatable field-service coordination workflows."),
    ("remodel", 14.0, "Remodeling markets create repeatable field-service coordination workflows."),
    ("repair", 14.0, "Repair markets create recurring service and dispatch workflows."),
    ("maintenance", 12.0, "Maintenance markets create recurring service and dispatch workflows."),
    ("collection", 8.0, "Collection workflows are route-based and operationally repetitive."),
]

NEGATIVE_TITLE_SIGNALS: list[tuple[str, float, str]] = [
    ("software", -28.0, "Software-native markets are not the intended first wedge."),
    ("programming", -24.0, "Custom software markets are not the intended first wedge."),
    ("computer systems design", -20.0, "IT-services markets are less aligned with the operator-heavy wedge thesis."),
    ("data processing", -22.0, "Data-processing markets are less aligned with the operator-heavy wedge thesis."),
    ("web hosting", -22.0, "Web-hosting markets are less aligned with the operator-heavy wedge thesis."),
    ("computing infrastructure", -24.0, "Infrastructure software markets are less aligned with the operator-heavy wedge thesis."),
    ("portfolio management", -26.0, "Capital-markets services are not the intended first wedge."),
    ("investment", -24.0, "Capital-markets services are not the intended first wedge."),
    ("securities", -24.0, "Capital-markets services are not the intended first wedge."),
    ("banking", -20.0, "Banking markets are institution-heavy and not the intended first wedge."),
    ("credit unions", -16.0, "Credit-union markets are institution-heavy and not the intended first wedge."),
    ("insurance carriers", -22.0, "Insurance-carrier markets are institution-heavy and not the intended first wedge."),
    ("insurance", -14.0, "Insurance buyers are often institution-heavy for a first wedge."),
    ("research and development", -18.0, "R&D services skew toward specialist knowledge-work rather than repeatable operator workflows."),
    ("universities", -18.0, "University buyers are institution-heavy and slower to wedge."),
    ("college", -18.0, "University buyers are institution-heavy and slower to wedge."),
    ("schools", -14.0, "School buyers are institution-heavy and slower to wedge."),
    ("museum", -16.0, "Museum markets are less consistent with the intended first-wedge profile."),
    ("golf", -14.0, "Country-club markets are less consistent with the intended first-wedge profile."),
    ("amusement", -18.0, "Amusement markets are less consistent with the intended first-wedge profile."),
    ("motion picture", -18.0, "Media-production markets are less aligned with the intended first wedge."),
    ("electric power", -16.0, "Utility infrastructure is less practical as a first wedge."),
    ("pipeline", -16.0, "Pipeline infrastructure is less practical as a first wedge."),
    ("air transportation", -16.0, "Air transportation is capital-intensive and less practical as a first wedge."),
    ("hospitals", -12.0, "Large institutional hospital systems are harder first-wedge buyers than local operator markets."),
    ("medical insurance", -18.0, "Payer markets are institution-heavy and not the intended first wedge."),
    ("pharmaceutical", -12.0, "Pharma manufacturing is less aligned with the intended first wedge."),
]


@dataclass(frozen=True)
class SourceSpec:
    artifact_id: str
    name: str
    url: str
    local_path: Path
    vintage: str
    description: str


@dataclass(frozen=True)
class CBPRow2017:
    naics_code_2017: str
    naics_title_2017: str
    establishments: int
    employment: int
    annual_payroll_usd: int


@dataclass(frozen=True)
class BLSRow2022:
    naics_code_2022: str
    establishments: int
    employment: int
    average_annual_pay_usd: int
    employment_growth_pct: float
    pay_growth_pct: float
    disclosure_code: str


@dataclass(frozen=True)
class SBARow2022:
    naics_code_2022: str
    naics_title_2022: str
    size_standard_basis: str
    size_standard_value: float
    size_standard_display: str
    footnotes: list[str]


@dataclass(frozen=True)
class CrosswalkRow:
    naics_code_2022: str
    naics_title_2022: str
    status_code: str
    naics_codes_2017: list[str]
    naics_titles_2017: list[str]


@dataclass(frozen=True)
class MatrixIndustryOption:
    matrix_code: str
    matrix_title: str
    code_prefix: str


@dataclass(frozen=True)
class MatrixOccupationRow:
    occupation_title: str
    occupation_code: str
    occupation_type: str
    employment_2024_thousands: float
    percent_of_industry_2024: float
    percent_of_occupation_2024: float
    projected_employment_2034_thousands: float
    projected_percent_of_industry_2034: float
    projected_percent_of_occupation_2034: float
    employment_change_2024_2034_thousands: float
    employment_pct_change_2024_2034: float
    display_level: int


SOURCE_SPECS = [
    SourceSpec(
        artifact_id="cbp_2022_us_national",
        name="Census County Business Patterns 2022 national industry API pull",
        url=CBP_URL,
        local_path=CBP_RAW_PATH,
        vintage="2022",
        description="National business-demography anchors from CBP, still keyed on NAICS 2017.",
    ),
    SourceSpec(
        artifact_id="sba_size_standards_2023",
        name="SBA Table of Size Standards effective March 17, 2023",
        url=SBA_URL,
        local_path=SBA_RAW_PATH,
        vintage="2023-03-17",
        description="Small-business size thresholds keyed on NAICS 2022.",
    ),
    SourceSpec(
        artifact_id="bls_qcew_2024_us000",
        name="BLS QCEW 2024 annual national area slice",
        url=BLS_URL,
        local_path=BLS_RAW_PATH,
        vintage="2024",
        description="Private-sector annual employment and pay anchors keyed on NAICS 2022.",
    ),
    SourceSpec(
        artifact_id="naics_2022_to_2017_changes_only",
        name="Census 2022 to 2017 NAICS changes workbook",
        url=NAICS_CROSSWALK_URL,
        local_path=NAICS_CROSSWALK_RAW_PATH,
        vintage="2022",
        description="Official changes-only bridge used to normalize CBP 2017 rows into 2022 industries.",
    ),
    SourceSpec(
        artifact_id="bls_nem_2024_2034_industry_home",
        name="BLS National Employment Matrix industry directory page",
        url=BLS_MATRIX_HOME_URL,
        local_path=BLS_MATRIX_HOME_RAW_PATH,
        vintage="2024-2034",
        description="Official BLS matrix industry directory used to resolve available industry occupation profiles.",
    ),
    SourceSpec(
        artifact_id="onet_30_2_occupation_data",
        name="O*NET 30.2 occupation data text file",
        url=ONET_OCCUPATION_DATA_URL,
        local_path=ONET_OCCUPATION_DATA_RAW_PATH,
        vintage="30.2",
        description="O*NET occupation roster used to resolve preferred base O*NET-SOC rows for each SOC code.",
    ),
    SourceSpec(
        artifact_id="onet_30_2_work_activities",
        name="O*NET 30.2 work activities text file",
        url=ONET_WORK_ACTIVITIES_URL,
        local_path=ONET_WORK_ACTIVITIES_RAW_PATH,
        vintage="30.2",
        description="Occupation-level O*NET work-activity ratings used to score workflow burden from industry occupation mix.",
    ),
]


def build_first_slice(refresh: bool = False) -> dict[str, Any]:
    ensure_directories()

    fetched_sources = fetch_sources(refresh=refresh)
    cbp_2017 = load_cbp_rows(CBP_RAW_PATH)
    bls_2022 = load_bls_rows(BLS_RAW_PATH)
    sba_2022 = load_sba_rows(SBA_RAW_PATH)
    crosswalk_rows, new_to_old, old_to_new = load_crosswalk_rows(NAICS_CROSSWALK_RAW_PATH)
    all_candidate_codes = canonical_candidate_codes(cbp_2017, bls_2022, sba_2022, new_to_old, old_to_new)
    eligible_codes, base_coverage_gaps = split_eligible_codes(all_candidate_codes, bls_2022, sba_2022)

    write_csv(
        CLEAN_CROSSWALK_PATH,
        [
            {
                "naics_code_2022": row.naics_code_2022,
                "naics_title_2022": row.naics_title_2022,
                "status_code": row.status_code,
                "naics_codes_2017": "|".join(row.naics_codes_2017),
                "naics_titles_2017": "|".join(row.naics_titles_2017),
            }
            for row in crosswalk_rows
        ],
    )

    write_csv(
        CLEAN_CBP_2017_PATH,
        [
            {
                "naics_code_2017": row.naics_code_2017,
                "naics_title_2017": row.naics_title_2017,
                "establishments_2022": row.establishments,
                "employment_2022": row.employment,
                "annual_payroll_usd_2022": row.annual_payroll_usd,
                "average_annual_pay_usd_2022": safe_ratio(row.annual_payroll_usd, row.employment),
            }
            for row in cbp_2017.values()
        ],
    )

    cbp_2022_rows, cbp_coverage_gaps = normalize_cbp_to_2022(
        eligible_codes,
        cbp_2017,
        sba_2022,
        bls_2022,
        new_to_old,
        old_to_new,
    )
    coverage_gaps = merge_coverage_gaps(base_coverage_gaps, cbp_coverage_gaps)

    write_csv(
        CLEAN_CBP_2022_PATH,
        [
            {
                "naics_code_2022": row["naics_code_2022"],
                "naics_title_2022": row["naics_title_2022"],
                "cbp_mapping_type": row["cbp_mapping_type"],
                "cbp_source_naics_2017_codes": "|".join(row["cbp_source_naics_2017_codes"]),
                "cbp_establishments_2022": row["cbp_establishments_2022"],
                "cbp_employment_2022": row["cbp_employment_2022"],
                "cbp_annual_payroll_usd_2022": row["cbp_annual_payroll_usd_2022"],
                "cbp_average_annual_pay_usd_2022": row["cbp_average_annual_pay_usd_2022"],
            }
            for row in cbp_2022_rows.values()
        ],
    )

    write_csv(
        CLEAN_BLS_PATH,
        [
            {
                "naics_code_2022": row.naics_code_2022,
                "bls_establishments_2024": row.establishments,
                "bls_employment_2024": row.employment,
                "bls_average_annual_pay_usd_2024": row.average_annual_pay_usd,
                "bls_employment_growth_pct_2024": row.employment_growth_pct,
                "bls_pay_growth_pct_2024": row.pay_growth_pct,
                "bls_disclosure_code": row.disclosure_code,
            }
            for row in bls_2022.values()
        ],
    )

    write_csv(
        CLEAN_SBA_PATH,
        [
            {
                "naics_code_2022": row.naics_code_2022,
                "naics_title_2022": row.naics_title_2022,
                "size_standard_basis": row.size_standard_basis,
                "size_standard_value": row.size_standard_value,
                "size_standard_display": row.size_standard_display,
                "footnotes": "|".join(row.footnotes),
            }
            for row in sba_2022.values()
        ],
    )

    matrix_options = load_bls_matrix_options(BLS_MATRIX_HOME_RAW_PATH)
    write_csv(
        CLEAN_BLS_MATRIX_DIRECTORY_PATH,
        [
            {
                "matrix_code": option.matrix_code,
                "matrix_title": option.matrix_title,
                "code_prefix": option.code_prefix,
            }
            for option in matrix_options
        ],
    )
    preferred_onet_codes = load_preferred_onet_codes(ONET_OCCUPATION_DATA_RAW_PATH)
    onet_activity_profiles = load_onet_work_activity_profiles(
        ONET_WORK_ACTIVITIES_RAW_PATH,
        preferred_onet_codes,
    )

    full_rows = assemble_full_rows(cbp_2022_rows, bls_2022, sba_2022, coverage_gaps)
    attach_workflow_profiles(
        full_rows,
        matrix_options,
        onet_activity_profiles,
        refresh=refresh,
    )
    scored_rows = score_rows(full_rows)

    write_csv(
        CLEAN_INDUSTRY_TABLE_PATH,
        [
            flatten_for_csv(row)
            for row in scored_rows
        ],
    )
    write_csv(
        DATA_CELLS_CSV_PATH,
        [
            flatten_for_csv(row)
            for row in scored_rows
        ],
    )
    write_csv(
        CLEAN_WORKFLOW_PROFILE_PATH,
        [
            flatten_workflow_profile_for_csv(row)
            for row in scored_rows
        ],
    )
    write_csv(
        DATA_COVERAGE_GAPS_PATH,
        coverage_gap_rows(coverage_gaps),
    )

    artifacts = [build_source_artifact(spec, fetched_sources[spec.artifact_id]) for spec in SOURCE_SPECS]
    artifacts.append(
        build_local_artifact(
            artifact_id="bls_nem_2024_2034_industry_profiles",
            name="BLS National Employment Matrix industry occupation profiles",
            url=BLS_MATRIX_HOME_URL,
            local_path=BLS_MATRIX_PROFILES_RAW_PATH,
            vintage="2024-2034",
            description="Cached industry-by-occupation profiles expanded from the official BLS National Employment Matrix industry query surface.",
        )
    )
    payload = build_site_payload(scored_rows, coverage_gaps, artifacts)

    DATA_CELLS_JSON_PATH.write_text(json.dumps(scored_rows, indent=2) + "\n", encoding="utf-8")
    SITE_DATA_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")

    return payload


def ensure_directories() -> None:
    for directory in (RAW_DIR, CLEAN_DIR, DATA_DIR, SITE_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def fetch_sources(refresh: bool = False) -> dict[str, Path]:
    fetched: dict[str, Path] = {}
    for spec in SOURCE_SPECS:
        if refresh or not spec.local_path.exists():
            content = fetch_bytes(spec.url)
            spec.local_path.write_bytes(content)
        fetched[spec.artifact_id] = spec.local_path
    return fetched


def fetch_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request) as response:
        return response.read()


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", errors="ignore")


def load_cbp_rows(path: Path) -> dict[str, CBPRow2017]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    results: dict[str, CBPRow2017] = {}
    for naics_code, naics_title, establishments, employment, payann_k, *_ in rows[1:]:
        if not NUMERIC_CODE_RE.fullmatch(naics_code):
            continue
        annual_payroll_usd = int(payann_k) * 1_000
        results[naics_code] = CBPRow2017(
            naics_code_2017=naics_code,
            naics_title_2017=naics_title,
            establishments=int(establishments),
            employment=int(employment),
            annual_payroll_usd=annual_payroll_usd,
        )
    return results


def load_bls_rows(path: Path) -> dict[str, BLSRow2022]:
    rows: dict[str, BLSRow2022] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            code = raw["industry_code"].strip()
            if not NUMERIC_CODE_RE.fullmatch(code):
                continue
            if raw["own_code"] != "5" or raw["agglvl_code"] != "18":
                continue
            rows[code] = BLSRow2022(
                naics_code_2022=code,
                establishments=parse_int(raw["annual_avg_estabs"]),
                employment=parse_int(raw["annual_avg_emplvl"]),
                average_annual_pay_usd=parse_int(raw["avg_annual_pay"]),
                employment_growth_pct=parse_float(raw["oty_annual_avg_emplvl_pct_chg"]),
                pay_growth_pct=parse_float(raw["oty_avg_annual_pay_pct_chg"]),
                disclosure_code=raw["disclosure_code"].strip(),
            )
    return rows


def load_sba_rows(path: Path) -> dict[str, SBARow2022]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook["table_of_size_standards-all"]

    rows: dict[str, SBARow2022] = {}
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        code_value = raw[0]
        if not isinstance(code_value, int):
            continue
        code = f"{code_value:06d}"
        if not NUMERIC_CODE_RE.fullmatch(code):
            continue

        title = clean_sba_title(raw[1] or "")
        receipts = raw[2]
        employees = raw[3]
        footnotes = parse_footnotes(raw[4])

        if isinstance(receipts, (int, float)):
            basis = "receipts_millions_usd"
            value = float(receipts)
            display = f"${value:,.2f}M"
        elif isinstance(receipts, str) and "assets" in receipts.lower():
            basis = "assets_millions_usd"
            value = parse_first_number(receipts)
            display = compact_whitespace(receipts)
        elif employees not in ("", None):
            basis = "employees"
            value = float(employees)
            display = f"{int(value):,} employees"
        else:
            continue

        rows[code] = SBARow2022(
            naics_code_2022=code,
            naics_title_2022=title,
            size_standard_basis=basis,
            size_standard_value=value,
            size_standard_display=display,
            footnotes=footnotes,
        )
    return rows


def load_crosswalk_rows(path: Path) -> tuple[list[CrosswalkRow], dict[str, list[str]], dict[str, list[str]]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[CrosswalkRow] = []
    new_to_old: dict[str, list[str]] = {}
    old_to_new: defaultdict[str, list[str]] = defaultdict(list)

    for raw in sheet.iter_rows(min_row=3, values_only=True):
        code_2022, title_2022, status_code, codes_2017_raw, titles_2017_raw, *_ = raw
        if code_2022 is None or codes_2017_raw is None:
            continue

        new_code = f"{int(code_2022):06d}"
        old_codes = [
            piece.strip().lstrip("*")
            for piece in str(codes_2017_raw).replace("\n", "|").split("|")
            if piece and piece.strip()
        ]
        old_titles = [
            compact_whitespace(piece)
            for piece in str(titles_2017_raw).split("\n")
            if piece and compact_whitespace(piece)
        ]

        row = CrosswalkRow(
            naics_code_2022=new_code,
            naics_title_2022=compact_whitespace(title_2022 or ""),
            status_code=str(status_code or "").strip(),
            naics_codes_2017=old_codes,
            naics_titles_2017=old_titles,
        )
        rows.append(row)
        new_to_old[new_code] = old_codes
        for old_code in old_codes:
            old_to_new[old_code].append(new_code)

    return rows, new_to_old, dict(old_to_new)


def load_bls_matrix_options(path: Path) -> list[MatrixIndustryOption]:
    page = path.read_text(encoding="utf-8")
    quote = chr(34)
    pattern = re.compile(
        rf"<option value={quote}([^{quote}]+){quote}>([^<]+)",
        flags=re.IGNORECASE,
    )
    options: list[MatrixIndustryOption] = []
    for matrix_code, matrix_title in pattern.findall(page):
        title = html.unescape(compact_whitespace(matrix_title))
        options.append(
            MatrixIndustryOption(
                matrix_code=matrix_code,
                matrix_title=title,
                code_prefix=matrix_option_prefix(matrix_code),
            )
        )
    return options


def load_preferred_onet_codes(path: Path) -> dict[str, list[str]]:
    base_to_codes: defaultdict[str, list[str]] = defaultdict(list)
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for raw in reader:
            code = compact_whitespace(raw["O*NET-SOC Code"])
            if "." not in code:
                continue
            base_to_codes[code.split(".", 1)[0]].append(code)

    preferred: dict[str, list[str]] = {}
    for base_code, codes in base_to_codes.items():
        primary_code = f"{base_code}.00"
        if primary_code in codes:
            preferred[base_code] = [primary_code]
        else:
            preferred[base_code] = sorted(codes)
    return preferred


def load_onet_work_activity_profiles(
    path: Path,
    preferred_onet_codes: dict[str, list[str]],
) -> dict[str, dict[str, float]]:
    tracked_elements = {
        element_name
        for component in WORKFLOW_COMPONENT_WEIGHTS.values()
        for element_name, _ in component
    }
    activity_rows: defaultdict[str, dict[str, float]] = defaultdict(dict)
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for raw in reader:
            if raw["Scale ID"] != "IM":
                continue
            element_name = compact_whitespace(raw["Element Name"])
            if element_name not in tracked_elements:
                continue
            activity_rows[compact_whitespace(raw["O*NET-SOC Code"])][element_name] = parse_float(
                raw["Data Value"]
            )

    profiles: dict[str, dict[str, float]] = {}
    for base_code, codes in preferred_onet_codes.items():
        active_codes = [code for code in codes if code in activity_rows]
        if not active_codes:
            continue
        profiles[base_code] = {}
        for element_name in tracked_elements:
            values = [
                activity_rows[code][element_name]
                for code in active_codes
                if element_name in activity_rows[code]
            ]
            if values:
                profiles[base_code][element_name] = sum(values) / len(values)
    return profiles


def normalize_cbp_to_2022(
    candidate_codes: list[str],
    cbp_2017: dict[str, CBPRow2017],
    sba_2022: dict[str, SBARow2022],
    bls_2022: dict[str, BLSRow2022],
    new_to_old: dict[str, list[str]],
    old_to_new: dict[str, list[str]],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    coverage_gaps: dict[str, list[str]] = defaultdict(list)
    normalized: dict[str, dict[str, Any]] = {}

    for code in candidate_codes:
        supported = False
        source_codes: list[str] = []
        mapping_type = ""
        establishments = 0
        employment = 0
        annual_payroll_usd = 0

        if code in cbp_2017 and code not in old_to_new:
            supported = True
            mapping_type = "direct"
            source_codes = [code]
            source_row = cbp_2017[code]
            establishments = source_row.establishments
            employment = source_row.employment
            annual_payroll_usd = source_row.annual_payroll_usd
        elif code in new_to_old:
            source_codes = sorted(new_to_old[code])
            if not all(source_code in cbp_2017 for source_code in source_codes):
                coverage_gaps["missing_cbp_source_codes"].append(code)
                continue
            if any(len(old_to_new.get(source_code, [])) != 1 for source_code in source_codes):
                coverage_gaps["non_allocable_2017_split"].append(code)
                continue

            supported = True
            mapping_type = "aggregated_2017_to_2022"
            for source_code in source_codes:
                source_row = cbp_2017[source_code]
                establishments += source_row.establishments
                employment += source_row.employment
                annual_payroll_usd += source_row.annual_payroll_usd
        else:
            coverage_gaps["no_cbp_mapping"].append(code)
            continue

        if not supported:
            coverage_gaps["unsupported_cbp_mapping"].append(code)
            continue

        normalized[code] = {
            "naics_code_2022": code,
            "naics_title_2022": sba_2022[code].naics_title_2022,
            "cbp_mapping_type": mapping_type,
            "cbp_source_naics_2017_codes": source_codes,
            "cbp_establishments_2022": establishments,
            "cbp_employment_2022": employment,
            "cbp_annual_payroll_usd_2022": annual_payroll_usd,
            "cbp_average_annual_pay_usd_2022": round(safe_ratio(annual_payroll_usd, employment), 2),
        }

    for reason in sorted(coverage_gaps):
        coverage_gaps[reason] = sorted(set(coverage_gaps[reason]))

    return normalized, dict(coverage_gaps)


def assemble_full_rows(
    cbp_2022_rows: dict[str, dict[str, Any]],
    bls_2022: dict[str, BLSRow2022],
    sba_2022: dict[str, SBARow2022],
    coverage_gaps: dict[str, list[str]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for code in sorted(set(cbp_2022_rows) & set(bls_2022) & set(sba_2022)):
        cbp_row = cbp_2022_rows[code]
        bls_row = bls_2022[code]
        sba_row = sba_2022[code]

        employees_per_establishment = safe_ratio(
            cbp_row["cbp_employment_2022"],
            cbp_row["cbp_establishments_2022"],
        )
        payroll_per_establishment = safe_ratio(
            cbp_row["cbp_annual_payroll_usd_2022"],
            cbp_row["cbp_establishments_2022"],
        )
        employment_gap_ratio = safe_ratio(bls_row.employment, cbp_row["cbp_employment_2022"])

        caveats = [
            "National-only slice; regional density and local regulatory variation are not yet represented.",
            "Workflow burden is inferred from public industry structure and occupation mix, not company-level software workflow evidence.",
        ]
        if cbp_row["cbp_mapping_type"] == "aggregated_2017_to_2022":
            caveats.append(
                "CBP business anchors are aggregated forward from 2017 NAICS into the 2022 definition."
            )
        if employment_gap_ratio < 0.7 or employment_gap_ratio > 1.3:
            caveats.append(
                "CBP 2022 employment and BLS 2024 private-sector employment differ materially; treat labor scale as directional."
            )

        row = {
            "entity_id": f"naics:{code}",
            "entity_name": sba_row.naics_title_2022,
            "entity_type": "naics_industry",
            "naics_code": code,
            "naics_version": "2022",
            "geography": {
                "level": "national",
                "code": "US",
                "name": "United States",
            },
            "size_band": "all",
            "placeholder": False,
            "lineage": {
                "cbp_mapping_type": cbp_row["cbp_mapping_type"],
                "cbp_source_naics_2017_codes": cbp_row["cbp_source_naics_2017_codes"],
                "source_vintages": {
                    "cbp": "2022",
                    "sba": "2023-03-17",
                    "bls": "2024",
                },
            },
            "anchors": {
                "cbp_establishments": cbp_row["cbp_establishments_2022"],
                "cbp_employment": cbp_row["cbp_employment_2022"],
                "cbp_annual_payroll_usd": cbp_row["cbp_annual_payroll_usd_2022"],
                "cbp_average_annual_pay_usd": cbp_row["cbp_average_annual_pay_usd_2022"],
                "bls_establishments": bls_row.establishments,
                "bls_employment": bls_row.employment,
                "bls_average_annual_pay_usd": bls_row.average_annual_pay_usd,
                "bls_employment_growth_pct": bls_row.employment_growth_pct,
                "bls_pay_growth_pct": bls_row.pay_growth_pct,
                "sba_size_standard": {
                    "basis": sba_row.size_standard_basis,
                    "value": sba_row.size_standard_value,
                    "display": sba_row.size_standard_display,
                    "footnotes": sba_row.footnotes,
                },
            },
            "score_inputs": {
                "employees_per_establishment": round(employees_per_establishment, 2),
                "payroll_per_establishment_usd": round(payroll_per_establishment, 2),
                "employment_gap_ratio_bls_to_cbp": round(employment_gap_ratio, 3),
                "sba_size_standard_basis": sba_row.size_standard_basis,
                "sba_size_standard_value": sba_row.size_standard_value,
                "sector_code": code[:2],
                "sector_adjustment_note": SECTOR_FIT_ADJUSTMENTS.get(code[:2], (0.0, "No explicit sector-level fit adjustment."))[1],
                "thesis_fit_positive_signals": [],
                "thesis_fit_negative_signals": [],
            },
            "workflow_profile": {},
            "scores": {},
            "recommended_move": "",
            "summary": "",
            "evidence": [],
            "caveats": caveats,
            "sources": [
                "cbp_2022_us_national",
                "sba_size_standards_2023",
                "bls_qcew_2024_us000",
                "naics_2022_to_2017_changes_only",
            ],
        }
        rows.append(row)

    if coverage_gaps:
        rows.sort(key=lambda item: item["naics_code"])
    return rows


def attach_workflow_profiles(
    rows: list[dict[str, Any]],
    matrix_options: list[MatrixIndustryOption],
    onet_activity_profiles: dict[str, dict[str, float]],
    refresh: bool,
) -> None:
    if not rows:
        return

    candidate_mappings = {
        row["naics_code"]: candidate_matrix_industries(
            row["naics_code"],
            row["entity_name"],
            matrix_options,
        )
        for row in rows
    }
    matrix_codes = {
        mapping["matrix_code"]
        for mappings in candidate_mappings.values()
        for mapping in mappings
    }
    matrix_profiles = fetch_bls_matrix_profiles(matrix_codes, refresh=refresh)

    for row in rows:
        selected_mapping = None
        workflow_profile = None
        for index, mapping in enumerate(candidate_mappings[row["naics_code"]]):
            try:
                workflow_profile = build_industry_workflow_profile(
                    mapping,
                    matrix_profiles[mapping["matrix_code"]],
                    onet_activity_profiles,
                )
                selected_mapping = dict(mapping)
                if index > 0:
                    selected_mapping["mapping_note"] = (
                        selected_mapping["mapping_note"]
                        + " Fallback selected after a more specific BLS line did not expose detailed occupation rows."
                    )
                    workflow_profile["mapping_note"] = selected_mapping["mapping_note"]
                break
            except ValueError:
                continue
        if workflow_profile is None or selected_mapping is None:
            raise ValueError(
                f"No usable BLS workflow profile found for {row['naics_code']} {row['entity_name']!r}"
            )

        mapping = selected_mapping
        row["workflow_profile"] = workflow_profile
        row["sources"] = dedupe_preserve_order(
            row["sources"]
            + [
                "bls_nem_2024_2034_industry_home",
                "bls_nem_2024_2034_industry_profiles",
                "onet_30_2_occupation_data",
                "onet_30_2_work_activities",
            ]
        )
        if mapping["mapping_type"] == "numeric_parent":
            row["caveats"].append(
                "Workflow layer uses the closest published BLS parent industry because the matrix does not expose this NAICS line separately."
            )
        elif mapping["mapping_type"] == "broad_parent":
            row["caveats"].append(
                "Workflow layer uses a broader published BLS parent industry; treat occupation mix as directional rather than exact for this NAICS line."
            )
        if workflow_profile["occupation_coverage_share_pct"] < 85:
            row["caveats"].append(
                "Workflow score is based on the visible BLS occupation mix and may miss some small or suppressed occupations."
            )
        row["caveats"] = dedupe_preserve_order(row["caveats"])


def candidate_matrix_industries(
    naics_code: str,
    entity_name: str,
    matrix_options: list[MatrixIndustryOption],
) -> list[dict[str, str]]:
    by_code = {option.matrix_code: option for option in matrix_options}
    by_title: defaultdict[str, list[MatrixIndustryOption]] = defaultdict(list)
    for option in matrix_options:
        by_title[normalize_lookup_text(option.matrix_title)].append(option)

    candidates: list[dict[str, str]] = []
    seen_codes: set[str] = set()

    def add_candidate(option: MatrixIndustryOption, mapping_type: str, mapping_note: str) -> None:
        if option.matrix_code in seen_codes:
            return
        seen_codes.add(option.matrix_code)
        candidates.append(
            {
                "matrix_code": option.matrix_code,
                "matrix_title": option.matrix_title,
                "mapping_type": mapping_type,
                "mapping_note": mapping_note,
            }
        )

    if naics_code in by_code:
        option = by_code[naics_code]
        add_candidate(
            option,
            "exact_code",
            "BLS matrix publishes this exact industry code.",
        )

    title_matches = by_title.get(normalize_lookup_text(entity_name), [])
    if len(title_matches) == 1:
        option = title_matches[0]
        add_candidate(
            option,
            "exact_title",
            "BLS matrix publishes a title-equivalent industry, but not the exact NAICS code.",
        )

    parent_candidates: list[tuple[int, tuple[int, int], str, MatrixIndustryOption]] = []
    for option in matrix_options:
        if len(option.code_prefix) < 3:
            continue
        if naics_code.startswith(option.code_prefix):
            parent_candidates.append(
                (
                    len(option.code_prefix),
                    title_token_score(entity_name, option.matrix_title),
                    option.matrix_code,
                    option,
                )
            )

    if not parent_candidates:
        raise ValueError(f"No BLS matrix industry mapping found for {naics_code} {entity_name!r}")

    parent_candidates.sort(reverse=True)
    for prefix_length, _, _, option in parent_candidates:
        mapping_type = "numeric_parent" if prefix_length >= 4 else "broad_parent"
        mapping_note = (
            "BLS matrix rolls this industry into a closely related parent line."
            if mapping_type == "numeric_parent"
            else "BLS matrix only publishes a broader parent line for this industry."
        )
        add_candidate(option, mapping_type, mapping_note)

    return candidates


def fetch_bls_matrix_profiles(
    matrix_codes: set[str],
    refresh: bool,
) -> dict[str, dict[str, Any]]:
    existing_profiles: dict[str, dict[str, Any]] = {}
    if not refresh and BLS_MATRIX_PROFILES_RAW_PATH.exists():
        raw_payload = json.loads(BLS_MATRIX_PROFILES_RAW_PATH.read_text(encoding="utf-8"))
        existing_profiles = raw_payload.get("profiles", {})

    codes_to_fetch = sorted(matrix_codes) if refresh else sorted(matrix_codes - set(existing_profiles))
    if codes_to_fetch:
        with ThreadPoolExecutor(max_workers=8) as executor:
            for matrix_code, profile in executor.map(fetch_single_bls_matrix_profile, codes_to_fetch):
                existing_profiles[matrix_code] = profile
        BLS_MATRIX_PROFILES_RAW_PATH.write_text(
            json.dumps(
                {
                    "generated_at": iso_timestamp(),
                    "source_url": BLS_MATRIX_HOME_URL,
                    "profiles": {
                        code: existing_profiles[code]
                        for code in sorted(existing_profiles)
                    },
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    return existing_profiles


def fetch_single_bls_matrix_profile(matrix_code: str) -> tuple[str, dict[str, Any]]:
    html_text = fetch_text(BLS_MATRIX_QUERY_URL_TEMPLATE.format(code=matrix_code))
    return matrix_code, parse_bls_matrix_profile_page(matrix_code, html_text)


def parse_bls_matrix_profile_page(matrix_code: str, page: str) -> dict[str, Any]:
    title_match = re.search(
        rf"<strong>{re.escape(matrix_code)}\s+([^<]+)</strong>",
        page,
        flags=re.IGNORECASE,
    )
    matrix_title = html.unescape(compact_whitespace(title_match.group(1))) if title_match else matrix_code

    tbody_match = re.search(r"<tbody>(.*?)</tbody>", page, flags=re.IGNORECASE | re.DOTALL)
    if not tbody_match:
        raise ValueError(f"Could not parse occupation table for matrix code {matrix_code}")

    rows: list[dict[str, Any]] = []
    for row_html in re.findall(r"<tr>(.*?)</tr>", tbody_match.group(1), flags=re.IGNORECASE | re.DOTALL):
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_html, flags=re.IGNORECASE | re.DOTALL)
        if len(cells) != 13:
            continue
        cleaned = [
            compact_whitespace(html.unescape(re.sub(r"<[^>]+>", " ", cell)))
            for cell in cells
        ]
        rows.append(
            {
                "occupation_title": cleaned[0],
                "occupation_code": cleaned[1],
                "occupation_type": cleaned[2],
                "employment_2024_thousands": parse_numeric_cell(cleaned[3]),
                "percent_of_industry_2024": parse_numeric_cell(cleaned[4]),
                "percent_of_occupation_2024": parse_numeric_cell(cleaned[5]),
                "projected_employment_2034_thousands": parse_numeric_cell(cleaned[6]),
                "projected_percent_of_industry_2034": parse_numeric_cell(cleaned[7]),
                "projected_percent_of_occupation_2034": parse_numeric_cell(cleaned[8]),
                "employment_change_2024_2034_thousands": parse_numeric_cell(cleaned[9]),
                "employment_pct_change_2024_2034": parse_numeric_cell(cleaned[10]),
                "occupation_sort": cleaned[11],
                "display_level": int(cleaned[12] or 0),
            }
        )

    return {
        "matrix_code": matrix_code,
        "matrix_title": matrix_title,
        "rows": rows,
    }


def build_industry_workflow_profile(
    mapping: dict[str, str],
    matrix_profile: dict[str, Any],
    onet_activity_profiles: dict[str, dict[str, float]],
) -> dict[str, Any]:
    line_item_rows = [
        MatrixOccupationRow(
            occupation_title=raw["occupation_title"],
            occupation_code=raw["occupation_code"],
            occupation_type=raw["occupation_type"],
            employment_2024_thousands=raw["employment_2024_thousands"],
            percent_of_industry_2024=raw["percent_of_industry_2024"],
            percent_of_occupation_2024=raw["percent_of_occupation_2024"],
            projected_employment_2034_thousands=raw["projected_employment_2034_thousands"],
            projected_percent_of_industry_2034=raw["projected_percent_of_industry_2034"],
            projected_percent_of_occupation_2034=raw["projected_percent_of_occupation_2034"],
            employment_change_2024_2034_thousands=raw["employment_change_2024_2034_thousands"],
            employment_pct_change_2024_2034=raw["employment_pct_change_2024_2034"],
            display_level=raw["display_level"],
        )
        for raw in matrix_profile["rows"]
        if raw["occupation_type"] == "Line Item" and SOC_CODE_RE.fullmatch(raw["occupation_code"])
    ]
    if not line_item_rows:
        raise ValueError(f"No line-item occupations found for matrix code {mapping['matrix_code']}")

    weighted_rows: list[tuple[float, MatrixOccupationRow, dict[str, float]]] = []
    covered_share = 0.0
    for occupation in line_item_rows:
        onet_profile = onet_activity_profiles.get(occupation.occupation_code)
        if not onet_profile:
            continue
        share = occupation.percent_of_industry_2024
        if share <= 0:
            continue
        weighted_rows.append((share, occupation, onet_profile))
        covered_share += share

    if not weighted_rows or covered_share <= 0:
        raise ValueError(f"No O*NET workflow coverage found for matrix code {mapping['matrix_code']}")

    component_scores = {
        component_name: round(
            compute_weighted_workflow_component(weighted_rows, element_weights),
            1,
        )
        for component_name, element_weights in WORKFLOW_COMPONENT_WEIGHTS.items()
    }
    component_blend = weighted_average(
        [
            (WORKFLOW_SCORE_WEIGHTS[component_name], component_scores[component_name])
            for component_name in WORKFLOW_SCORE_WEIGHTS
        ]
    )
    frontline_operator_share_pct = (
        sum(
            share
            for share, occupation, _ in weighted_rows
            if soc_major_group(occupation.occupation_code) in FRONTLINE_OPERATOR_MAJOR_GROUPS
        )
        / covered_share
        * 100.0
    )
    knowledge_work_share_pct = (
        sum(
            share
            for share, occupation, _ in weighted_rows
            if soc_major_group(occupation.occupation_code) in KNOWLEDGE_WORK_MAJOR_GROUPS
        )
        / covered_share
        * 100.0
    )
    workflow_intensity = weighted_average(
        [
            (0.45, component_blend),
            (0.35, frontline_operator_share_pct),
            (0.20, 100.0 - knowledge_work_share_pct),
        ]
    )
    top_occupations = [
        {
            "occupation_code": occupation.occupation_code,
            "occupation_title": occupation.occupation_title,
            "employment_2024_thousands": round(occupation.employment_2024_thousands, 1),
            "percent_of_industry": round(occupation.percent_of_industry_2024, 1),
        }
        for _, occupation, _ in sorted(
            weighted_rows,
            key=lambda item: (-item[0], item[1].occupation_title),
        )[:5]
    ]

    return {
        "matrix_industry_code": mapping["matrix_code"],
        "matrix_industry_title": mapping["matrix_title"],
        "mapping_type": mapping["mapping_type"],
        "mapping_note": mapping["mapping_note"],
        "occupation_coverage_share_pct": round(covered_share, 1),
        "frontline_operator_share_pct": round(frontline_operator_share_pct, 1),
        "knowledge_work_share_pct": round(knowledge_work_share_pct, 1),
        "workflow_intensity": round(workflow_intensity, 1),
        "component_scores": component_scores,
        "top_occupations": top_occupations,
    }


def compute_weighted_workflow_component(
    weighted_rows: list[tuple[float, MatrixOccupationRow, dict[str, float]]],
    element_weights: list[tuple[str, float]],
) -> float:
    total_share = sum(share for share, _, _ in weighted_rows)
    if total_share <= 0:
        return 0.0

    weighted_component = 0.0
    for share, _, onet_profile in weighted_rows:
        component_value = compute_activity_component_score(onet_profile, element_weights)
        weighted_component += (share / total_share) * component_value
    return weighted_component


def compute_activity_component_score(
    onet_profile: dict[str, float],
    element_weights: list[tuple[str, float]],
) -> float:
    weighted_total = 0.0
    total_weight = 0.0
    for element_name, weight in element_weights:
        if element_name not in onet_profile:
            continue
        weighted_total += weight * onet_profile[element_name]
        total_weight += weight
    if total_weight <= 0:
        return 0.0
    return (weighted_total / total_weight) * 20.0


def score_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return rows

    estab_logs = [math.log10(row["anchors"]["cbp_establishments"] + 1) for row in rows]
    emp_per_estab_logs = [
        math.log10(row["score_inputs"]["employees_per_establishment"] + 1) for row in rows
    ]
    payroll_per_estab_logs = [
        math.log10(row["score_inputs"]["payroll_per_establishment_usd"] + 1) for row in rows
    ]
    pay_logs = [math.log10(row["anchors"]["bls_average_annual_pay_usd"] + 1) for row in rows]
    growth_values = [row["anchors"]["bls_employment_growth_pct"] for row in rows]
    pay_growth_values = [row["anchors"]["bls_pay_growth_pct"] for row in rows]
    employment_logs = [math.log10(row["anchors"]["cbp_employment"] + 1) for row in rows]
    payroll_logs = [math.log10(row["anchors"]["cbp_annual_payroll_usd"] + 1) for row in rows]
    dollar_threshold_logs = [
        math.log10(
            row["anchors"]["sba_size_standard"]["value"] * 1_000_000 + 1
        )
        for row in rows
        if row["anchors"]["sba_size_standard"]["basis"] in {"receipts_millions_usd", "assets_millions_usd"}
    ]
    employee_threshold_logs = [
        math.log10(row["anchors"]["sba_size_standard"]["value"] + 1)
        for row in rows
        if row["anchors"]["sba_size_standard"]["basis"] == "employees"
    ]

    for row in rows:
        estab_pct = percentile(estab_logs, math.log10(row["anchors"]["cbp_establishments"] + 1))
        small_firm_pct = inverse_percentile(
            emp_per_estab_logs,
            math.log10(row["score_inputs"]["employees_per_establishment"] + 1),
        )
        employees_per_estab_pct = percentile(
            emp_per_estab_logs,
            math.log10(row["score_inputs"]["employees_per_establishment"] + 1),
        )
        payroll_per_estab_pct = percentile(
            payroll_per_estab_logs,
            math.log10(row["score_inputs"]["payroll_per_establishment_usd"] + 1),
        )
        pay_pct = percentile(
            pay_logs,
            math.log10(row["anchors"]["bls_average_annual_pay_usd"] + 1),
        )
        growth_pct = percentile(growth_values, row["anchors"]["bls_employment_growth_pct"])
        pay_growth_pct = percentile(pay_growth_values, row["anchors"]["bls_pay_growth_pct"])
        employment_pct = percentile(
            employment_logs,
            math.log10(row["anchors"]["cbp_employment"] + 1),
        )
        payroll_pct = percentile(
            payroll_logs,
            math.log10(row["anchors"]["cbp_annual_payroll_usd"] + 1),
        )

        sba_basis = row["anchors"]["sba_size_standard"]["basis"]
        if sba_basis in {"receipts_millions_usd", "assets_millions_usd"}:
            sba_pct = percentile(
                dollar_threshold_logs,
                math.log10(row["anchors"]["sba_size_standard"]["value"] * 1_000_000 + 1),
            )
        else:
            sba_pct = percentile(
                employee_threshold_logs,
                math.log10(row["anchors"]["sba_size_standard"]["value"] + 1),
            )

        fragmentation = weighted_average(
            [(0.55, estab_pct), (0.45, small_firm_pct)]
        )
        inverse_pay_pct = inverse_percentile(
            pay_logs,
            math.log10(row["anchors"]["bls_average_annual_pay_usd"] + 1),
        )
        operating_complexity = weighted_average(
            [(0.7, employees_per_estab_pct), (0.3, inverse_pay_pct)]
        )
        willingness_to_pay = weighted_average(
            [(0.55, sba_pct), (0.45, payroll_per_estab_pct)]
        )
        growth = weighted_average(
            [(0.7, growth_pct), (0.3, pay_growth_pct)]
        )
        market_scale = weighted_average(
            [(0.5, employment_pct), (0.5, payroll_pct)]
        )
        workflow_intensity = row["workflow_profile"]["workflow_intensity"]
        thesis_fit, positive_signals, negative_signals = compute_thesis_fit(
            row,
            estab_pct=estab_pct,
            pay_pct=pay_pct,
        )
        row["score_inputs"]["thesis_fit_positive_signals"] = positive_signals
        row["score_inputs"]["thesis_fit_negative_signals"] = negative_signals

        software_wedge = weighted_average(
            [
                (0.24, fragmentation),
                (0.2, operating_complexity),
                (0.16, workflow_intensity),
                (0.14, willingness_to_pay),
                (0.1, growth),
                (0.04, market_scale),
                (0.12, thesis_fit),
            ]
        )
        rollup_wedge = weighted_average(
            [
                (0.33, fragmentation),
                (0.18, market_scale),
                (0.18, sba_pct),
                (0.08, operating_complexity),
                (0.08, workflow_intensity),
                (0.05, growth),
                (0.1, thesis_fit),
            ]
        )

        confidence = 90.0
        if row["lineage"]["cbp_mapping_type"] == "aggregated_2017_to_2022":
            confidence -= 10.0
        employment_gap_ratio = row["score_inputs"]["employment_gap_ratio_bls_to_cbp"]
        if employment_gap_ratio < 0.7 or employment_gap_ratio > 1.3:
            confidence -= 5.0

        row["scores"] = {
            "fragmentation": round(fragmentation, 1),
            "operating_complexity": round(operating_complexity, 1),
            "willingness_to_pay": round(willingness_to_pay, 1),
            "growth": round(growth, 1),
            "market_scale": round(market_scale, 1),
            "workflow_intensity": round(workflow_intensity, 1),
            "thesis_fit": round(thesis_fit, 1),
            "software_wedge": round(software_wedge, 1),
            "rollup_wedge": round(rollup_wedge, 1),
            "confidence": round(max(confidence, 55.0), 1),
        }
        row["recommended_move"] = recommend_move(row["scores"])
        row["summary"] = build_summary(row)
        row["evidence"] = build_evidence(row)

    rows.sort(
        key=lambda item: (
            move_priority(item["recommended_move"]),
            -primary_rank_score(item),
            -item["scores"]["software_wedge"],
            -item["scores"]["rollup_wedge"],
            item["entity_name"],
        )
    )
    for index, row in enumerate(rows, start=1):
        row["rank"] = index

    return rows


def build_site_payload(
    scored_rows: list[dict[str, Any]],
    coverage_gaps: dict[str, list[str]],
    source_artifacts: list[dict[str, Any]],
) -> dict[str, Any]:
    move_counts = Counter(row["recommended_move"] for row in scored_rows)
    score_values = [row["scores"]["software_wedge"] for row in scored_rows]
    rollup_values = [row["scores"]["rollup_wedge"] for row in scored_rows]
    workflow_values = [row["scores"]["workflow_intensity"] for row in scored_rows]
    confidence_values = [row["scores"]["confidence"] for row in scored_rows]

    excluded_total = sum(len(values) for values in coverage_gaps.values())
    return {
        "generated_at": iso_timestamp(),
        "method_version": "first-slice-v3-workflow-intensity",
        "canonical_naics_version": "2022",
        "summary": {
            "fully_joined_rows": len(scored_rows),
            "excluded_rows": excluded_total,
            "excluded_by_reason": {
                reason: len(values)
                for reason, values in sorted(coverage_gaps.items())
            },
            "recommended_move_counts": dict(sorted(move_counts.items())),
            "software_wedge_range": [min(score_values), max(score_values)],
            "rollup_wedge_range": [min(rollup_values), max(rollup_values)],
            "workflow_intensity_range": [min(workflow_values), max(workflow_values)],
            "confidence_range": [min(confidence_values), max(confidence_values)],
            "national_only": True,
        },
        "coverage_gaps": {
            reason: {
                "count": len(values),
                "sample_codes": values[:15],
            }
            for reason, values in sorted(coverage_gaps.items())
        },
        "source_artifacts": source_artifacts,
        "filters": {
            "recommended_moves": sorted(move_counts),
            "score_fields": [
                "software_wedge",
                "rollup_wedge",
                "fragmentation",
                "operating_complexity",
                "willingness_to_pay",
                "growth",
                "market_scale",
                "workflow_intensity",
                "thesis_fit",
                "confidence",
            ],
        },
        "entities": scored_rows,
    }


def build_source_artifact(spec: SourceSpec, local_path: Path) -> dict[str, Any]:
    content = local_path.read_bytes()
    return {
        "artifact_id": spec.artifact_id,
        "name": spec.name,
        "url": spec.url,
        "local_path": str(local_path.relative_to(ROOT)),
        "vintage": spec.vintage,
        "description": spec.description,
        "sha256": hashlib.sha256(content).hexdigest(),
        "size_bytes": len(content),
    }


def build_local_artifact(
    artifact_id: str,
    name: str,
    url: str,
    local_path: Path,
    vintage: str,
    description: str,
) -> dict[str, Any]:
    content = local_path.read_bytes()
    return {
        "artifact_id": artifact_id,
        "name": name,
        "url": url,
        "local_path": str(local_path.relative_to(ROOT)),
        "vintage": vintage,
        "description": description,
        "sha256": hashlib.sha256(content).hexdigest(),
        "size_bytes": len(content),
    }


def build_evidence(row: dict[str, Any]) -> list[dict[str, Any]]:
    anchors = row["anchors"]
    score_inputs = row["score_inputs"]
    positive_signals = score_inputs["thesis_fit_positive_signals"]
    negative_signals = score_inputs["thesis_fit_negative_signals"]
    workflow_profile = row["workflow_profile"]
    evidence = [
        {
            "label": "CBP footprint",
            "detail": (
                f"{anchors['cbp_establishments']:,} establishments and "
                f"{anchors['cbp_employment']:,} employees in 2022."
            ),
            "source_ids": ["cbp_2022_us_national"],
        },
        {
            "label": "BLS labor pricing",
            "detail": (
                f"Private-sector average annual pay was "
                f"${anchors['bls_average_annual_pay_usd']:,.0f} in 2024."
            ),
            "source_ids": ["bls_qcew_2024_us000"],
        },
        {
            "label": "BLS momentum",
            "detail": (
                f"Employment changed {anchors['bls_employment_growth_pct']:+.1f}% "
                f"and pay changed {anchors['bls_pay_growth_pct']:+.1f}% year over year."
            ),
            "source_ids": ["bls_qcew_2024_us000"],
        },
        {
            "label": "Operator density",
            "detail": (
                f"Average establishment size is "
                f"{score_inputs['employees_per_establishment']:,.1f} employees."
            ),
            "source_ids": ["cbp_2022_us_national"],
        },
        {
            "label": "SBA buyer boundary",
            "detail": (
                f"SBA small-business threshold is "
                f"{anchors['sba_size_standard']['display']}."
            ),
            "source_ids": ["sba_size_standards_2023"],
        },
        {
            "label": "Occupation workflow mix",
            "detail": build_workflow_evidence_detail(workflow_profile),
            "source_ids": [
                "bls_nem_2024_2034_industry_profiles",
                "onet_30_2_work_activities",
            ],
        },
        {
            "label": "Thesis fit",
            "detail": build_thesis_fit_detail(positive_signals, negative_signals),
            "source_ids": [
                "cbp_2022_us_national",
                "sba_size_standards_2023",
                "bls_qcew_2024_us000",
                "bls_nem_2024_2034_industry_profiles",
                "onet_30_2_work_activities",
            ],
        },
    ]
    if row["lineage"]["cbp_mapping_type"] == "aggregated_2017_to_2022":
        evidence.append(
            {
                "label": "Crosswalk note",
                "detail": (
                    "CBP anchors aggregate the 2017 NAICS source codes "
                    + ", ".join(row["lineage"]["cbp_source_naics_2017_codes"])
                    + " into this 2022 definition."
                ),
                "source_ids": [
                    "cbp_2022_us_national",
                    "naics_2022_to_2017_changes_only",
                ],
            }
        )
    return evidence


def build_summary(row: dict[str, Any]) -> str:
    move = row["recommended_move"]
    scores = row["scores"]
    lead_positive = row["score_inputs"]["thesis_fit_positive_signals"][:1]
    lead_negative = row["score_inputs"]["thesis_fit_negative_signals"][:1]
    if move == "build":
        fit_phrase = lead_positive[0] if lead_positive else "the thesis-fit signals stay supportive"
        return (
            f"{fit_phrase} and the structure is fragmented enough to support a build-first wedge."
        )
    if move == "acquire":
        fit_phrase = lead_negative[0] if lead_negative else "the fit is better for fragmentation than for a clean software wedge"
        return (
            f"The market looks more compelling as a fragmented roll-up lane because {fit_phrase}."
        )
    if move == "sell":
        return (
            "The structure supports go-to-market attention now, but the software wedge still "
            "looks less decisive than the very top build candidates."
        )
    if move == "monitor":
        if scores["growth"] >= scores["willingness_to_pay"]:
            return "Momentum is present, but the buyer-quality signals are still mixed."
        return "There is some structural promise here, but the current evidence is not strong enough for a first move."
    return "The current public-data signals look too weak for a near-term build, sell, or acquire motion."


def recommend_move(scores: dict[str, float]) -> str:
    if (
        scores["software_wedge"] >= 60
        and scores["fragmentation"] >= 65
        and scores["operating_complexity"] >= 40
        and scores["workflow_intensity"] >= 55
        and scores["willingness_to_pay"] >= 35
        and scores["thesis_fit"] >= 60
    ):
        return "build"
    if (
        scores["rollup_wedge"] >= 68
        and scores["fragmentation"] >= 75
        and scores["workflow_intensity"] >= 40
        and scores["thesis_fit"] >= 40
    ):
        return "acquire"
    if scores["software_wedge"] >= 58 or scores["rollup_wedge"] >= 60:
        return "sell"
    if scores["software_wedge"] >= 45 or scores["growth"] >= 55:
        return "monitor"
    return "ignore"


def coverage_gap_rows(coverage_gaps: dict[str, list[str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for reason, codes in sorted(coverage_gaps.items()):
        for code in codes:
            rows.append(
                {
                    "naics_code_2022": code,
                    "reason": reason,
                }
            )
    return rows


def flatten_for_csv(row: dict[str, Any]) -> dict[str, Any]:
    anchors = row["anchors"]
    sba_size = anchors["sba_size_standard"]
    workflow_profile = row["workflow_profile"]
    return {
        "rank": row["rank"],
        "naics_code": row["naics_code"],
        "entity_name": row["entity_name"],
        "cbp_mapping_type": row["lineage"]["cbp_mapping_type"],
        "cbp_source_naics_2017_codes": "|".join(row["lineage"]["cbp_source_naics_2017_codes"]),
        "cbp_establishments_2022": anchors["cbp_establishments"],
        "cbp_employment_2022": anchors["cbp_employment"],
        "cbp_annual_payroll_usd_2022": anchors["cbp_annual_payroll_usd"],
        "cbp_average_annual_pay_usd_2022": anchors["cbp_average_annual_pay_usd"],
        "bls_establishments_2024": anchors["bls_establishments"],
        "bls_employment_2024": anchors["bls_employment"],
        "bls_average_annual_pay_usd_2024": anchors["bls_average_annual_pay_usd"],
        "bls_employment_growth_pct_2024": anchors["bls_employment_growth_pct"],
        "bls_pay_growth_pct_2024": anchors["bls_pay_growth_pct"],
        "sba_size_standard_basis": sba_size["basis"],
        "sba_size_standard_value": sba_size["value"],
        "sba_size_standard_display": sba_size["display"],
        "employees_per_establishment": row["score_inputs"]["employees_per_establishment"],
        "payroll_per_establishment_usd": row["score_inputs"]["payroll_per_establishment_usd"],
        "employment_gap_ratio_bls_to_cbp": row["score_inputs"]["employment_gap_ratio_bls_to_cbp"],
        "fragmentation": row["scores"]["fragmentation"],
        "operating_complexity": row["scores"]["operating_complexity"],
        "willingness_to_pay": row["scores"]["willingness_to_pay"],
        "growth": row["scores"]["growth"],
        "market_scale": row["scores"]["market_scale"],
        "workflow_intensity": row["scores"]["workflow_intensity"],
        "thesis_fit": row["scores"]["thesis_fit"],
        "software_wedge": row["scores"]["software_wedge"],
        "rollup_wedge": row["scores"]["rollup_wedge"],
        "confidence": row["scores"]["confidence"],
        "workflow_matrix_industry_code": workflow_profile["matrix_industry_code"],
        "workflow_mapping_type": workflow_profile["mapping_type"],
        "workflow_coverage_share_pct": workflow_profile["occupation_coverage_share_pct"],
        "workflow_frontline_operator_share_pct": workflow_profile["frontline_operator_share_pct"],
        "workflow_knowledge_work_share_pct": workflow_profile["knowledge_work_share_pct"],
        "workflow_documentation": workflow_profile["component_scores"]["documentation"],
        "workflow_coordination": workflow_profile["component_scores"]["coordination"],
        "workflow_compliance": workflow_profile["component_scores"]["compliance"],
        "workflow_care_service": workflow_profile["component_scores"]["care_service"],
        "workflow_top_occupations": "|".join(
            f"{occupation['occupation_title']} ({occupation['percent_of_industry']:.1f}%)"
            for occupation in workflow_profile["top_occupations"][:3]
        ),
        "recommended_move": row["recommended_move"],
        "summary": row["summary"],
        "caveats": " | ".join(row["caveats"]),
    }


def flatten_workflow_profile_for_csv(row: dict[str, Any]) -> dict[str, Any]:
    workflow_profile = row["workflow_profile"]
    top_occupations = workflow_profile["top_occupations"]
    return {
        "rank": row["rank"],
        "naics_code": row["naics_code"],
        "entity_name": row["entity_name"],
        "matrix_industry_code": workflow_profile["matrix_industry_code"],
        "matrix_industry_title": workflow_profile["matrix_industry_title"],
        "mapping_type": workflow_profile["mapping_type"],
        "mapping_note": workflow_profile["mapping_note"],
        "occupation_coverage_share_pct": workflow_profile["occupation_coverage_share_pct"],
        "frontline_operator_share_pct": workflow_profile["frontline_operator_share_pct"],
        "knowledge_work_share_pct": workflow_profile["knowledge_work_share_pct"],
        "workflow_intensity": workflow_profile["workflow_intensity"],
        "documentation": workflow_profile["component_scores"]["documentation"],
        "coordination": workflow_profile["component_scores"]["coordination"],
        "compliance": workflow_profile["component_scores"]["compliance"],
        "care_service": workflow_profile["component_scores"]["care_service"],
        "top_occupation_1": format_workflow_occupation(top_occupations, 0),
        "top_occupation_2": format_workflow_occupation(top_occupations, 1),
        "top_occupation_3": format_workflow_occupation(top_occupations, 2),
    }


def parse_int(value: str) -> int:
    if value in ("", None):
        return 0
    return int(float(value))


def parse_float(value: str) -> float:
    if value in ("", None):
        return 0.0
    return float(value)


def parse_footnotes(value: Any) -> list[str]:
    if value in ("", None):
        return []
    return [compact_whitespace(piece) for piece in str(value).split(",") if compact_whitespace(piece)]


def percentile(values: list[float], value: float) -> float:
    ordered = sorted(values)
    count = len(ordered)
    if count == 0:
        return 50.0
    lower = sum(1 for item in ordered if item < value)
    equal = sum(1 for item in ordered if item == value)
    return round(((lower + (equal / 2)) / count) * 100, 1)


def inverse_percentile(values: list[float], value: float) -> float:
    return round(100.0 - percentile(values, value), 1)


def weighted_average(weighted_values: list[tuple[float, float]]) -> float:
    numerator = sum(weight * value for weight, value in weighted_values)
    denominator = sum(weight for weight, _ in weighted_values)
    if denominator == 0:
        return 0.0
    return numerator / denominator


def safe_ratio(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    return numerator / denominator


def compact_whitespace(value: str) -> str:
    return " ".join(str(value).split())


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def compute_thesis_fit(row: dict[str, Any], estab_pct: float, pay_pct: float) -> tuple[float, list[str], list[str]]:
    title = row["entity_name"].lower()
    sector_code = row["score_inputs"]["sector_code"]
    employees_per_establishment = row["score_inputs"]["employees_per_establishment"]
    establishments = row["anchors"]["cbp_establishments"]
    workflow_profile = row["workflow_profile"]
    workflow_components = workflow_profile["component_scores"]
    workflow_intensity = workflow_profile["workflow_intensity"]

    score = 50.0
    positive_signals: list[str] = []
    negative_signals: list[str] = []

    sector_adjustment, sector_note = SECTOR_FIT_ADJUSTMENTS.get(
        sector_code,
        (0.0, "No explicit sector-level fit adjustment."),
    )
    score += sector_adjustment
    if sector_adjustment > 0:
        positive_signals.append(sector_note)
    elif sector_adjustment < 0:
        negative_signals.append(sector_note)

    for pattern, adjustment, detail in POSITIVE_TITLE_SIGNALS:
        if pattern in title:
            score += adjustment
            positive_signals.append(detail)

    for pattern, adjustment, detail in NEGATIVE_TITLE_SIGNALS:
        if pattern in title:
            score += adjustment
            negative_signals.append(detail)

    if establishments >= 10_000:
        score += 6.0
        positive_signals.append("The market already has a large national base of operating locations.")
    elif establishments <= 500:
        score -= 4.0
        negative_signals.append("The market has a relatively small national establishment base for a first wedge.")

    if 5 <= employees_per_establishment <= 80:
        score += 8.0
        positive_signals.append("Average site size suggests many distributed operating units rather than a few giant locations.")
    elif employees_per_establishment > 250:
        score -= 8.0
        negative_signals.append("Very large average site size can point toward enterprise selling rather than a distributed operator wedge.")
    elif employees_per_establishment < 3:
        score -= 6.0
        negative_signals.append("Very small average site size can make the workflow wedge thinner at the company level.")

    if pay_pct >= 85:
        score -= 8.0
        negative_signals.append("An elite wage profile often signals knowledge-work or institution-heavy buying behavior.")
    elif pay_pct <= 35:
        score += 4.0
        positive_signals.append("A moderate wage profile is more consistent with labor-operations software than elite knowledge-work tooling.")

    if estab_pct >= 70:
        score += 4.0
        positive_signals.append("The market is already broad enough nationally to support a real first go-to-market wedge.")

    if workflow_intensity >= 70:
        score += 8.0
        positive_signals.append("The occupation mix carries strong repeatable workflow load across documentation, coordination, or compliance.")
    elif workflow_intensity >= 60:
        score += 4.0
        positive_signals.append("The occupation mix shows real day-to-day workflow burden.")
    elif workflow_intensity <= 35:
        score -= 6.0
        negative_signals.append("The occupation mix looks lighter on repeatable workflow burden than the strongest operator wedges.")

    if workflow_components["documentation"] >= 65:
        positive_signals.append("The occupation mix leans heavily on documentation and recordkeeping work.")
    if workflow_components["coordination"] >= 65:
        positive_signals.append("The occupation mix requires frequent scheduling and worker coordination.")
    if workflow_components["compliance"] >= 60:
        positive_signals.append("The occupation mix carries meaningful monitoring and standards/compliance work.")
    if workflow_components["care_service"] >= 60:
        positive_signals.append("Frontline care and service work adds handoff and service-complexity pressure.")

    positive_signals = dedupe_preserve_order(positive_signals)
    negative_signals = dedupe_preserve_order(negative_signals)
    return clamp(score, 0.0, 100.0), positive_signals[:5], negative_signals[:5]


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def build_thesis_fit_detail(positive_signals: list[str], negative_signals: list[str]) -> str:
    if positive_signals and negative_signals:
        return (
            f"Positive fit signals: {positive_signals[0]} "
            f"Counter-signal: {negative_signals[0]}"
        )
    if positive_signals:
        return f"Positive fit signals: {positive_signals[0]}"
    if negative_signals:
        return f"Counter-signal: {negative_signals[0]}"
    return "No strong positive or negative thesis-fit signal surfaced beyond the structural data."


def build_workflow_evidence_detail(workflow_profile: dict[str, Any]) -> str:
    top_occupation = workflow_profile["top_occupations"][0]
    components = workflow_profile["component_scores"]
    return (
        f"Mapped to BLS matrix industry {workflow_profile['matrix_industry_code']} "
        f"({workflow_profile['matrix_industry_title']}); top visible occupation is "
        f"{top_occupation['occupation_title']} at {top_occupation['percent_of_industry']:.1f}% of industry employment. "
        f"Workflow intensity is {workflow_profile['workflow_intensity']:.1f}/100, with "
        f"documentation {components['documentation']:.1f}, "
        f"coordination {components['coordination']:.1f}, and "
        f"compliance {components['compliance']:.1f}. "
        f"Frontline operator share is {workflow_profile['frontline_operator_share_pct']:.1f}% of the visible occupation mix, "
        f"versus {workflow_profile['knowledge_work_share_pct']:.1f}% knowledge-work share."
    )


def format_workflow_occupation(top_occupations: list[dict[str, Any]], index: int) -> str:
    if index >= len(top_occupations):
        return ""
    occupation = top_occupations[index]
    return f"{occupation['occupation_title']} ({occupation['percent_of_industry']:.1f}%)"


def normalize_lookup_text(value: str) -> str:
    normalized = compact_whitespace(value).lower().replace("except", "")
    normalized = normalized.replace("‑", "-")
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def title_token_score(left: str, right: str) -> tuple[int, int]:
    left_tokens = set(normalize_lookup_text(left).split())
    right_tokens = set(normalize_lookup_text(right).split())
    overlap = len(left_tokens & right_tokens)
    spread_penalty = abs(len(left_tokens) - len(right_tokens))
    return (overlap, -spread_penalty)


def matrix_option_prefix(code: str) -> str:
    match = re.match(r"(\d+)", code)
    if not match:
        return ""
    return match.group(1).rstrip("0")


def parse_numeric_cell(value: str) -> float:
    if value in ("", None):
        return 0.0
    return float(str(value).replace(",", ""))


def soc_major_group(code: str) -> str:
    return code.split("-", 1)[0]


def canonical_candidate_codes(
    cbp_2017: dict[str, CBPRow2017],
    bls_2022: dict[str, BLSRow2022],
    sba_2022: dict[str, SBARow2022],
    new_to_old: dict[str, list[str]],
    old_to_new: dict[str, list[str]],
) -> list[str]:
    unchanged_cbp_codes = {code for code in cbp_2017 if code not in old_to_new}
    return sorted(set(bls_2022) | set(sba_2022) | set(new_to_old) | unchanged_cbp_codes)


def split_eligible_codes(
    all_candidate_codes: list[str],
    bls_2022: dict[str, BLSRow2022],
    sba_2022: dict[str, SBARow2022],
) -> tuple[list[str], dict[str, list[str]]]:
    eligible: list[str] = []
    coverage_gaps: dict[str, list[str]] = defaultdict(list)

    for code in all_candidate_codes:
        has_bls = code in bls_2022
        has_sba = code in sba_2022
        if has_bls and has_sba:
            eligible.append(code)
        elif has_bls and not has_sba:
            coverage_gaps["missing_sba_row"].append(code)
        elif has_sba and not has_bls:
            coverage_gaps["missing_bls_row"].append(code)
        else:
            coverage_gaps["missing_bls_and_sba_row"].append(code)

    return eligible, dict(coverage_gaps)


def merge_coverage_gaps(*gap_sets: dict[str, list[str]]) -> dict[str, list[str]]:
    merged: dict[str, set[str]] = defaultdict(set)
    for gap_set in gap_sets:
        for reason, codes in gap_set.items():
            merged[reason].update(codes)
    return {reason: sorted(codes) for reason, codes in sorted(merged.items())}


def move_priority(move: str) -> int:
    return {
        "build": 0,
        "acquire": 1,
        "sell": 2,
        "monitor": 3,
        "ignore": 4,
    }[move]


def primary_rank_score(row: dict[str, Any]) -> float:
    if row["recommended_move"] == "build":
        return row["scores"]["software_wedge"]
    if row["recommended_move"] == "acquire":
        return row["scores"]["rollup_wedge"]
    return max(row["scores"]["software_wedge"], row["scores"]["rollup_wedge"])


def clean_sba_title(value: str) -> str:
    return re.sub(r"(?<=\D)\d+$", "", compact_whitespace(value)).strip()


def parse_first_number(value: str) -> float:
    match = re.search(r"(\d+(?:\.\d+)?)", str(value))
    if not match:
        raise ValueError(f"Could not parse numeric value from {value!r}")
    return float(match.group(1))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def iso_timestamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()
